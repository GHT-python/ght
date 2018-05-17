import MySQLdb
from openpyxl import Workbook
import time
import winreg

conn=MySQLdb.connect(host='192.168.1.184',port=3306,user='dev_read',passwd='frt4d2*qdk',db='wzxy3_0_pro_page',charset='utf8')
# conn=MySQLdb.connect(host='192.168.202.204',port=30055,user='dev_read',passwd='frt4d2*qdk',db='wzxy3_0_pro_page',charset='utf8')

cursor = conn.cursor()
localtime = time.strftime('%Y-%m-%d',time.localtime(time.time()))

date ='\''+localtime+'%'+'\''

# date_start ="'2010-04-01 00:00:00'"
# date_end ="'2018-05-10 00:00:00'"
date_start = '\''+input("请输入开始时间：（格式如：2018-01-01 07:00:00）")+'\''
date_end = '\''+input("请输入结束时间：（格式如：2018-02-01 07:00:00）")+'\''
d1 = date_start[6:8]
d2 = date_start[9:11]
d3 = date_end[6:8]
d4 = date_end[9:11]

sql1 = '''SELECT b.name 学校名字,IFNULL(a.number,0) 学校公告数
FROM
(select B.name,B.id FROM  ninespring_foundation_schoolstat A, ninespring_foundation_schools B
WHERE A.school_id=B.id and A.created_at like %s and A.teacher_staff_count!=0 and B.salesman!='林文斌') b 
left join 
(select count(*) number,school_id from ninespring_schoolnotice_list where ninespring_schoolnotice_list.created_at between %s and %s 
and ninespring_schoolnotice_list.deleted_at is null GROUP BY school_id) a 
on a.school_id=b.id ORDER BY b.name''' % (date,date_start,date_end)
sql2 = '''SELECT b.name 学校名字,IFNULL(a.number,0) 班级公告数 FROM
(select B.name,B.id FROM  ninespring_foundation_schoolstat A, ninespring_foundation_schools B
WHERE A.school_id=B.id and A.created_at like %s and A.teacher_staff_count!=0 and B.salesman!='林文斌') b 
LEFT JOIN
(SELECT b.school_id,count(*) number FROM ninespring_classnotice_list a
LEFT JOIN ninespring_foundation_classes b ON a.class_id = b.id
where a.created_at between %s and %s and a.deleted_at is null GROUP BY b.school_id) a 
on a.school_id=b.id ORDER BY b.name''' % (date,date_start,date_end)
sql3 = '''SELECT b.name,IFNULL(a.number,0) 班级作业使用量
FROM
(select B.name,B.id FROM  ninespring_foundation_schoolstat A, ninespring_foundation_schools B
WHERE A.school_id=B.id and A.created_at like %s and A.teacher_staff_count!=0 and B.salesman!='林文斌') b 
LEFT JOIN
(SELECT b.school_id,count(*) number FROM ninespring_homework_task a
INNER JOIN ninespring_foundation_classes b ON b.id = a.class_id
where a.created_at between %s and %s and a.deleted_at is null GROUP BY b.school_id) a
on a.school_id=b.id ORDER BY b.name''' % (date,date_start,date_end)
sql4 = '''SELECT b.name 学校名字,IFNULL(a.number,0) 老师考勤使用量
FROM
(select B.name,B.id FROM  ninespring_foundation_schoolstat A, ninespring_foundation_schools B
WHERE A.school_id=B.id and A.created_at like %s and A.teacher_staff_count!=0 and B.salesman!='林文斌') b 
left join 
(select count(*) number,school_id from ninespring_attendance_record_details a where a.created_at between %s and %s
and a.deleted_at is null GROUP BY school_id) a 
on a.school_id=b.id ORDER BY b.name''' % (date,date_start,date_end)
sql5 = '''SELECT b.name 学校名字,IFNULL(a.number,0) 成长日记使用量
FROM
(select B.name,B.id FROM  ninespring_foundation_schoolstat A, ninespring_foundation_schools B
WHERE A.school_id=B.id and A.created_at like %s and A.teacher_staff_count!=0 and B.salesman!='林文斌') b 
LEFT JOIN
(SELECT school_id,count(*) number FROM ninespring_growth_logs a
INNER JOIN ninespring_foundation_parent_student_class_relations b ON a.student_id = b.student_id
INNER JOIN ninespring_foundation_classes c ON b.class_id = c.id
where a.created_at between %s and %s and  a.deleted_at is null GROUP BY school_id) a
on a.school_id=b.id ORDER BY b.name''' % (date,date_start,date_end)
sql6 = '''SELECT b.name 学校名字,IFNULL(a.number,0) 家长动态使用量
FROM
(select B.name,B.id FROM  ninespring_foundation_schoolstat A, ninespring_foundation_schools B
WHERE A.school_id=B.id and A.created_at like %s and A.teacher_staff_count!=0 and B.salesman!='林文斌') b 
left join 
(select a.school_id,count(b.id) number from ninespring_foundation_classes a inner join
(select b.class_id,a.id from ninespring_dynamic_index a
left join ninespring_foundation_parent_student_class_relations b on a.student_id=b.student_id
where a.created_at between %s and %s and a.deleted_at is null and a.student_id!=0) b on a.id=b.class_id GROUP BY a.school_id) a 
on a.school_id=b.id ORDER BY b.name''' % (date,date_start,date_end)
sql7 = '''SELECT b.name,IFNULL(a.number,0) 老师动态使用量 FROM
(select B.name,B.id FROM  ninespring_foundation_schoolstat A, ninespring_foundation_schools B
WHERE A.school_id=B.id and A.created_at like %s and A.teacher_staff_count!=0 and B.salesman!='林文斌') b 
left join 
(select count(*) number,b.school_id from ninespring_dynamic_index a
left join ninespring_foundation_classes b on a.class_id=b.id
where a.created_at between %s and %s and a.deleted_at is null and a.student_id=0 GROUP BY b.school_id) a 
on a.school_id=b.id ORDER BY b.name''' % (date,date_start,date_end)
sql8 = '''SELECT b.name 学校名字,IFNULL(a.number,0) 教师审批使用量 FROM
(select B.name,B.id FROM  ninespring_foundation_schoolstat A, ninespring_foundation_schools B
WHERE A.school_id=B.id and A.created_at like %s and A.teacher_staff_count!=0 and B.salesman!='林文斌') b 
left join 
(select count(*) number,school_id from ninespring_approve_log a where a.created_at between %s and %s 
and a.deleted_at is null GROUP BY school_id) a 
on a.school_id=b.id ORDER BY b.name''' % (date,date_start,date_end)
sql9 = '''SELECT b.name 学校名字,IFNULL(a.number,0) 手机端德育评价
FROM
(select B.name,B.id FROM  ninespring_foundation_schoolstat A, ninespring_foundation_schools B
WHERE A.school_id=B.id and A.created_at like %s and A.teacher_staff_count!=0 and B.salesman!='林文斌') b 
LEFT JOIN
(SELECT a.school_id,count(*) number FROM ninespring_eclasscard_moral_student_detail a 
where a.source_type=2 and a.created_at BETWEEN %s and %s and a.deleted_at is null GROUP BY school_id) a
on a.school_id=b.id ORDER BY b.name''' % (date,date_start,date_end)
sql10 = '''SELECT b.name 学校名字,IFNULL(a.number,0) 手机端成绩使用量
FROM
(select B.name,B.id FROM  ninespring_foundation_schoolstat A, ninespring_foundation_schools B
WHERE A.school_id=B.id and A.created_at like %s and A.teacher_staff_count!=0 and B.salesman!='林文斌') b 
left join 
(select count(*) number,a.school_id from ninespring_score_publish a where a.created_at 
between %s and %s and a.deleted_at is null GROUP BY a.school_id) a 
on a.school_id=b.id ORDER BY b.name''' % (date,date_start,date_end)
sql11 = '''SELECT b.name 学校名字,IFNULL(a.number,0) 后台成绩使用量
FROM
(select B.name,B.id FROM  ninespring_foundation_schoolstat A, ninespring_foundation_schools B
WHERE A.school_id=B.id and A.created_at like %s and A.teacher_staff_count!=0 and B.salesman!='林文斌') b 
left join 
(select count(*) number,a.school_id from ninespring_score_list a where a.created_at 
between %s and %s and a.deleted_at is null GROUP BY a.school_id) a 
on a.school_id=b.id ORDER BY b.name''' % (date,date_start,date_end)
sql12 = '''SELECT b.name 学校名字,IFNULL(a.number,0) 五卡使用量
FROM
(select B.name,B.id FROM  ninespring_foundation_schoolstat A, ninespring_foundation_schools B
WHERE A.school_id=B.id and A.created_at like %s and A.teacher_staff_count!=0 and B.salesman!='林文斌') b 
LEFT JOIN
(SELECT b.school_id,count(*) number FROM ninespring_performance_reward_card_index a
INNER JOIN ninespring_foundation_classes b ON a.class_id = b.id
where a.created_at between %s and %s and a.deleted_at is null GROUP BY b.school_id) a
on a.school_id=b.id ORDER BY b.name''' % (date,date_start,date_end)
sql13 = '''SELECT b.name 学校名字,IFNULL(a.number,0) 奖状使用量
FROM
(select B.name,B.id FROM  ninespring_foundation_schoolstat A, ninespring_foundation_schools B
WHERE A.school_id=B.id and A.created_at like %s and A.teacher_staff_count!=0 and B.salesman!='林文斌') b 
LEFT JOIN
(SELECT b.school_id,count(*) number FROM ninespring_performance_prize a
INNER JOIN ninespring_foundation_classes b ON a.class_id = b.id
where a.created_at between %s and %s and a.deleted_at is null GROUP BY b.school_id) a
on a.school_id=b.id ORDER BY b.name''' % (date,date_start,date_end)
sql14 = '''SELECT b.name 学校名字,IFNULL(a.number,0) 教师发起投票使用量
FROM
(select B.name,B.id FROM  ninespring_foundation_schoolstat A, ninespring_foundation_schools B
WHERE A.school_id=B.id and A.created_at like %s and A.teacher_staff_count!=0 and B.salesman!='林文斌') b 
LEFT JOIN
(SELECT school_id,count(*) number FROM  ninespring_vote_configs a
where a.created_at between %s and %s and a.deleted_at is null GROUP BY school_id) a
on a.school_id=b.id ORDER BY b.name''' % (date,date_start,date_end)
sql15 = '''SELECT b.name 学校名字,IFNULL(a.number,0) 家长投票使用量
FROM
(select B.name,B.id FROM  ninespring_foundation_schoolstat A, ninespring_foundation_schools B
WHERE A.school_id=B.id and A.created_at like %s and A.teacher_staff_count!=0 and B.salesman!='林文斌') b 
LEFT JOIN
(SELECT school_id,count(DISTINCT a.user_id) number FROM ninespring_vote_logs a
INNER JOIN ninespring_vote_configs b ON a.vote_config_id = b.id
where a.created_at between %s and %s and a.deleted_at is null GROUP BY school_id) a
on a.school_id=b.id ORDER BY b.name''' % (date,date_start,date_end)
sql16 = '''SELECT b.name 学校名字,IFNULL(a.number,0) 学生请假使用量
FROM
(select B.name,B.id FROM  ninespring_foundation_schoolstat A, ninespring_foundation_schools B
WHERE A.school_id=B.id and A.created_at like %s and A.teacher_staff_count!=0 and B.salesman!='林文斌') b 
LEFT JOIN
(SELECT school_id,count(*) number FROM ninespring_leave_application a
INNER JOIN ninespring_foundation_classes b ON a.class_id=b.id
where a.created_at between %s and %s and a.deleted_at is null GROUP BY school_id) a
on a.school_id=b.id ORDER BY b.name''' % (date,date_start,date_end)
sql17 = '''SELECT b.name 学校名字,IFNULL(a.number,0) 工资条使用量
FROM
(select B.name,B.id FROM  ninespring_foundation_schoolstat A, ninespring_foundation_schools B
WHERE A.school_id=B.id and A.created_at like %s and A.teacher_staff_count!=0 and B.salesman!='林文斌') b 
left join 
(select count(*) number,school_id from ninespring_salary_index a where a.created_at 
between %s and %s and a.deleted_at is null GROUP BY school_id) a 
on a.school_id=b.id ORDER BY b.name''' % (date,date_start,date_end)
sql18 = '''SELECT b.name 学校名字,IFNULL(a.number,0) 超级表单使用量 FROM 
(select B.name,B.id FROM  ninespring_foundation_schoolstat A, ninespring_foundation_schools B
WHERE A.school_id=B.id and A.created_at like %s and A.teacher_staff_count!=0 and B.salesman!='林文斌') b 
left join 
(SELECT school_id,count(*) number FROM ninespring_input_index a where a.created_at 
between %s and %s and a.deleted_at is null GROUP BY school_id) a 
on a.school_id=b.id ORDER BY b.name''' % (date,date_start,date_end)

cursor.execute(sql1)
result1 = cursor.fetchall()

cursor.execute(sql2)
result2 = cursor.fetchall()

cursor.execute(sql3)
result3 = cursor.fetchall()

cursor.execute(sql4)
result4 = cursor.fetchall()

cursor.execute(sql5)
result5 = cursor.fetchall()

cursor.execute(sql6)
result6 = cursor.fetchall()

cursor.execute(sql7)
result7 = cursor.fetchall()

cursor.execute(sql8)
result8 = cursor.fetchall()

cursor.execute(sql9)
result9 = cursor.fetchall()

cursor.execute(sql10)
result10 = cursor.fetchall()

cursor.execute(sql11)
result11 = cursor.fetchall()

cursor.execute(sql12)
result12 = cursor.fetchall()

cursor.execute(sql13)
result13 = cursor.fetchall()

cursor.execute(sql14)
result14 = cursor.fetchall()

cursor.execute(sql15)
result15 = cursor.fetchall()

cursor.execute(sql16)
result16 = cursor.fetchall()

cursor.execute(sql17)
result17 = cursor.fetchall()

cursor.execute(sql18)
result18 = cursor.fetchall()

# print(result1)
# print(result2)
# print(result3)
# print(result4)
# print(result5)
# print(result6)
# print(result7)
# print(result8)
# print(result9)
# print(result10)
# print(result11)
# print(result12)
# print(result13)
# print(result14)
# print(result15)
# print(result16)
# print(result17)
# print(result18)

cursor.close()
conn.close()
len = len(result1)
result = [result1,result2,result3,result4,result5,result6,result7,result8,result9,result10,result11,result12,result13,result14,result15,result16,result17,result18]
wb = Workbook()
def get_desktop():
    key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,r'Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders')
    return winreg.QueryValueEx(key, "Desktop")[0]
filename = r'C:\Users\guohaitao\Desktop'+'\功能使用数据'+d1+'.'+d2+'-'+d3+'.'+d4+'.xlsx'
fileds = [u'学校名称',u'学校公告数',u'班级公告数',u'班级作业使用量',u'老师考勤使用量',u'成长日记使用量',u'家长动态使用量',u'老师动态使用量',u'审批使用量',u'手机端德育评价',
          u'手机端成绩使用量',u'后台成绩使用量',u'五卡使用量',u'奖状使用量',u'老师发起投票量',u'家长投票使用量',u'学生请假使用量',u'工资条使用量',u'超级表单使用量']
ws = wb.active
ws.append(fileds)
for i in range(len):
    ws.cell(row=i+2,column=1,value=result1[i][0])
for i in range(18):
    for j in range(len):
        ws.cell(row=j+2,column=i+2,value=result[i][j][1])
wb.save(filename)
print ('查询完成')